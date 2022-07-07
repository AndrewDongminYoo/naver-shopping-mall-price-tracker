import os
import re
import csv
import requests
from datetime import datetime
from urllib.parse import quote, urlparse

from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from selenium.webdriver.chrome.service import Service
from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import sys

if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
    chrome_path = os.path.join(sys._MEIPASS, 'chromedriver.exe')
else:
    chrome_path = ChromeDriverManager().install()
chrome_service = Service(chrome_path)

chrome_options = Options()
chrome_options.add_argument("--no-sandbox")
# chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--remote-debugging-port=9230")
load_dotenv("prod.env")
filename = "크롤링 취합자료.xlsx"
sheet_name = "아이템 정보"

NAVER_CLIENT_ID = os.getenv('NAVER_CLIENT_ID')
NAVER_CLIENT_SECRET = os.getenv('NAVER_CLIENT_SECRET')

host_dict = dict(
    eleven_street="https://www.11st.co.kr/",
    lotte_i_mall="https://www.lotteimall.com/",
    lotte_on="https://www.lotteon.com/",
    g_market="http://item.gmarket.co.kr/",
    gee9="https://www.g9.co.kr/",
    gs_shop="https://with.gsshop.com/",
    on_style="https://display.cjonstyle.com/",
    auction="http://itempage3.auction.co.kr/",
    interpark="https://shopping.interpark.com/",
    smart_store="https://smartstore.naver.com/",
    we_make_price="https://front.wemakeprice.com/",
    tmon="https://www.tmon.co.kr/",
    unit="http://mahaknit.com/",
    shopping="https://shopping.naver.com/",
)

host_ko = {
    "https://www.11st.co.kr/": "11번가",
    "https://www.lotteimall.com/": "롯데아이몰",
    "https://www.lotteon.com/": "롯데온",
    "http://item.gmarket.co.kr/": "지마켓",
    "https://www.g9.co.kr/": "지구",
    "https://with.gsshop.com/": "지에스샵",
    "https://display.cjonstyle.com/": "온스타일",
    "http://itempage3.auction.co.kr/": "옥션",
    "https://shopping.interpark.com/": "인터파크",
    "https://smartstore.naver.com/": "스마트스토어",
    "https://front.wemakeprice.com/": "위메프",
    "https://www.tmon.co.kr/": "티몬",
    "http://mahaknit.com/": "마하유닛",
    "https://shopping.naver.com/": "네이버쇼핑",
}

product_type = {
    "1": "일반상품 가격비교 상품",
    "2": "가격비교 비매칭 일반상품",
    "3": "가격비교 매칭 일반상품",
    "4": "중고상품 가격비교 상품",
    "5": "가격비교 비매칭 중고상품",
    "6": "가격비교 매칭 중고상품",
    "7": "단종상품 가격비교 상품",
    "8": "가격비교 비매칭 단종상품",
    "9": "가격비교 매칭 단종상품",
    "10": "판매예정상품 가격비교 상품",
    "11": "가격비교 비매칭 판매예정상품",
    "12": "가격비교 매칭 판매예정상품",
}

store_types = {
    '02-772-3343': '본점', '02-2143-7251': '잠실점', '02-2164-5353': '영등포점', '02-3707-1321': '청량리점',
    '02-3289-8007': '관악점', '02-531-2224': '강남점', '02-950-2268': '노원점', '02-944-2274': '미아점',
    '02-2218-3404': '건대점', '02-6116-3051': '김포공항점', '02-6965-2661': '서울역점', '031-738-2217': '분당점',
    '031-909-3486': '일산점', '032-320-7298': '중동점', '031-412-7772': '안산점', '031-8086-9250': '평촌점',
    '031-8066-0286': '수원점', '032-242-2234': '인천터미널점', '031-8036-3746': '동탄점', '051-810-4280': '부산본점',
    '062-221-1308': '광주점', '042-601-2337': '대전점', '054-230-1345': '포항점', '052-960-4954': '울산점',
    '051-668-4254': '동래점', '055-279-3377': '창원점', '053-660-3322': '대구점', '053-258-3213': '상인점',
    '063-289-3252': '전주점', '061-801-2156': '남악점', '051-678-3488': '광복점'
}

store_types2 = {
    "EBJ": "본점", "IBJ": "본점", "EB": "본점", "IB": "본점", "BJ": "본점", "IJS": "잠실점", "EYD": "영등포점", "YDP": "영등포점",
    "ECL": "청량리점", "EGN": "강남점", "ENW": "노원점", "INW": "노원점", "TNW": "노원점", "EGP": "김포공항점", "EIS": "일산점", "ISS": "일산점",
    "EJD": "중동점", "IJD": "중동점", "EAS": "안산점", "IAS": "안산점", "AS": "안산점", "EPC": "평촌점", "IM": "평촌점", "ESW": "수원점",
    "ISW": "수원점", "ICC": "인천터미널점", "IC": "인천터미널점", "EDT": "동탄점", "IMT": "동탄점", "MDT": "동탄점", "EBS": "부산본점",
    "EGJ": "광주점", "IJ": "광주점", "EDJ": "대전점", "IDJ": "대전점", "EDR": "동래점", "IDR": "동래점", "EDG": "대구점", "IDG": "대구점",
    "DGG": "대구점", "EGB": "광복점",
}


def extract_title(string):
    string = re.sub(r"\[?유닛]?", "", string)
    string = re.sub(r"\[?롯데백화점( 2관)?]?", "", string)
    string = re.sub(r"\[?\(?AVE\)?]?", "", string)
    string = re.sub(r"</?b>", "", string)
    string = re.sub(r"\s+", " ", string)
    return string.strip()


def extract_phone(string):
    phone_regex = re.compile(r"(0\d{1,2}-\d{3,4}-\d{4})|(1\d{3}-\d{4})")
    match = phone_regex.search(string)
    if match:
        return match.group()
    return ""


def find_cs_number(page_source: str):
    phone = extract_phone(page_source)
    if phone in store_types:
        return store_types[phone]
    return phone


def get_host_from_url(url_string):
    hosts = list(host_dict.values())
    parsed = urlparse(url_string)
    host_url = f"{parsed.scheme}://{parsed.netloc}/"
    if host_url not in hosts:
        hosts.append(host_url)
    return host_url


def redirect_url(driver, url):
    find_all_href = [url]
    req = None
    while find_all_href:
        url = find_all_href[0]
        req = requests.get(url, allow_redirects=True)
        find_all_href = re.compile(r"https://cr.shopping.naver.com/[\-a-zA-Z\d%&?.=/#_~+()|]+").findall(req.text)
    find_all_redirects = re.compile(r'targetUrl = "([\-a-zA-Z\d:%&?.=/_#~+()|]+)"').findall(req.text)
    driver.get(find_all_redirects[0]) if find_all_redirects else driver.get(url)
    tabs = driver.window_handles
    if len(tabs) > 1:
        driver.switch_to.window(tabs.pop())
        driver.close()
        driver.switch_to.window(tabs[0])
    wait = WebDriverWait(driver, 30)
    wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "body > div")))
    return driver.page_source, driver.current_url


def find_model_name(source, word):
    model_name = re.search(f"{word}-([A-Za-z]+)", source)
    if model_name:
        if (model_name.group(1)).upper() in store_types2:
            return store_types2[model_name.group(1)]
    return ""


def scroll_infinite(webdriver):
    scroll_to_bottom = "window.scrollTo(0, document.body.scrollHeight);"
    get_window_height = "return document.body.scrollHeight"
    last_height = webdriver.execute_script(get_window_height)
    while True:
        webdriver.execute_script(scroll_to_bottom)
        import time
        time.sleep(0.5)
        new_height = webdriver.execute_script(get_window_height)
        if new_height == last_height:
            break
        last_height = new_height


def naver_shopping_search(csv_writer, index: int, season: str, code: str, low_price: int):
    display = 100
    start = 1

    url = f"https://openapi.naver.com/v1/search/shop?" \
          f"query={quote(code)}" \
          f"&display={str(display)}" \
          f"&start={str(start)}"
    headers = {
        "X-Naver-Client-Id": NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET
    }
    response = requests.get(url, headers=headers)
    start += display
    body = response.json()
    if "items" in body:
        with Chrome(service=chrome_service, options=chrome_options) as driver:
            driver.implicitly_wait(10)
            for data in body["items"]:
                if '의류' not in data['category1']:
                    print(f"{code}: 해당 코드로 패션의류 카테고리가 아닌 상품이 검색됩니다. 키워드를 확인하세요.")
                    continue
                elif not bigger_than(data['lprice'], low_price):
                    try:
                        title = extract_title(data['title'])
                        source, url = redirect_url(driver, data["link"])
                        cs_number = find_cs_number(source)
                        model_name = find_model_name(source, code)
                        host_key = get_host_from_url(url)
                        host_name = host_ko.get(host_key, host_key)
                        row = [
                            index, code, title, cs_number, model_name, season,
                            int(data["lprice"]), low_price,
                            host_name, url
                        ]
                        csv_writer.writerow(row)
                        print(row)
                    except Exception as e:
                        print(e)
    else:
        print(response.text)


def bigger_than(is_bigger: str, is_smaller: int):
    return int(is_bigger) >= is_smaller


def naver_without_api(csv_writer, index: int, season: str, code: str, korean_name: str, low_price: int):
    url = f"https://search.shopping.naver.com/search/all" \
          f"?catId=50000167" \
          f"&query={code}" \
          f"&pagingSize=80" \
          f"&sort=rel"
    with Chrome(service=chrome_service, options=chrome_options) as driver:
        driver.get(url)
        driver.implicitly_wait(10)
        scroll_infinite(driver)
        driver.find_elements(By.XPATH, '//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/ul/div/div/li')
        soup = BeautifulSoup(driver.page_source, "html.parser")
        for product in soup.select(
                '#__next > div > div.style_container__1YjHN > div > div.style_content_wrap__1PzEo > div.style_content__2T20F > ul > div > div > li'):
            lprice = product.select_one(
                'div > div.basicList_info_area__17Xyo > div.basicList_price_area__1UXXR > strong > span').text
            lprice = lprice.replace("최저", "").replace("원", "").replace(",", "")
            title_element = product.select_one(
                'div > div.basicList_info_area__17Xyo > div.basicList_title__3P9Q7 > a').text
            link_element = product.select_one(
                'div > div.basicList_info_area__17Xyo > div.basicList_title__3P9Q7 > a').get('href')
            if not bigger_than(lprice, low_price):
                try:
                    title = extract_title(title_element)
                    source, url = redirect_url(driver, link_element)
                    cs_number = find_cs_number(source)
                    model_name = find_model_name(source, code)
                    host_key = get_host_from_url(url)
                    host_name = host_ko.get(host_key, host_key)
                    row = [index, code, title, cs_number, model_name, season, int(lprice), low_price, host_name, url]
                    csv_writer.writerow(row)
                    print(row)
                except Exception as e:
                    print(e)


def main():
    dateformat = datetime.now().strftime("%Y%m%d-%H%M%S")
    new_filename = f"result_{dateformat}.csv"
    with open(new_filename, mode="w", newline="") as f:
        writer = csv.writer(f, delimiter=",", lineterminator="\n")
        header = ["NO", "스타일코드", "한글명", "지점_연락처", "지점_코드", "시즌", "판매가", "공식할인가", "판매처", "링크"]
        writer.writerow(header)
        wb: Workbook = load_workbook(
            filename=filename,
            data_only=True,
        )
        worksheet: Worksheet = wb[sheet_name]
        sheet_rows = worksheet.iter_rows(
            max_col=worksheet.max_column,
            max_row=worksheet.max_row,
            min_row=2,
            min_col=1,
            values_only=True,
        )
        for sheet_row in sheet_rows:
            index, code, korean_name, on_off, year, season, tag_price, dsc_price, percent = sheet_row
            # naver_shopping_search(writer, index, season, code, dsc_price)
            naver_without_api(writer, index, season, code, korean_name, dsc_price)
        f.close()


if __name__ == '__main__':
    main()
